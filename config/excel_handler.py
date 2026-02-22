from openpyxl import Workbook, load_workbook
from django.conf import settings
from django.http import HttpResponse
from rest_framework import status
from config.utils import errorcall, succescall, STATUS_ACTIVO
from django.db import transaction


class ExcelMasterHandler:
    def __init__(self, model, headers, filename_prefix, user_id):
        self.model = model
        self.headers = headers
        self.filename_prefix = filename_prefix
        self.user_id = user_id

    def generate_template(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"
        ws.append(self.headers)

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"))
        response["Content-Disposition"] = (
            f'attachment; filename="plantilla_{self.filename_prefix}.xlsx"'
        )
        wb.save(response)
        return response

    def export_data(self, queryset, serializer_class, field_mapping):
        """
        Exporta los datos usando un field_mapping:
        { 'CampoModelo': 'NombreColumnaExcel' }
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"

        # Cabeceras basadas en el orden del mapping
        headers = list(field_mapping.values())
        ws.append(headers)

        serializer = serializer_class(queryset, many=True)
        model_fields = list(field_mapping.keys())

        for item in serializer.data:
            row = [item.get(field, "") for field in model_fields]
            ws.append(row)

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"))
        response["Content-Disposition"] = (
            f'attachment; filename="{self.filename_prefix}_export.xlsx"'
        )
        wb.save(response)
        return response

    def import_data(
            self,
            request,
            validator_func,
            field_mapping,
            audit_save_fn=None):
        file = request.FILES.get("file")
        dry_run = request.data.get("dry_run", "false").lower() == "true"

        if not file:
            return errorcall(
                "No se ha proporcionado ningún archivo",
                status.HTTP_400_BAD_REQUEST,
            )

        # Validar tamaño máximo del archivo
        max_size = getattr(settings, "MAX_EXCEL_UPLOAD_SIZE", 5 * 1024 * 1024)
        if file.size > max_size:
            max_mb = max_size / (1024 * 1024)
            err_msg = (
                f"El archivo excede el tamaño máximo permitido "
                f"({max_mb:.0f} MB)"
            )
            return errorcall(
                err_msg,
                status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            )

        if not file.name.lower().endswith((".xlsx", ".xlsm")):
            return errorcall(
                "Solo se permiten archivos Excel (.xlsx, .xlsm)",
                status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                return errorcall(
                    "El archivo Excel está vacío", status.HTTP_400_BAD_REQUEST
                )

            headers_in_file = [str(h).strip() if h else "" for h in rows[0]]
            if not all(h in headers_in_file for h in self.headers):
                return errorcall(
                    f"Columnas requeridas faltantes. Se espera: "
                    f"{', '.join(self.headers)}",
                    status.HTTP_400_BAD_REQUEST,
                )

            # Mapeo de índices
            col_indices = {h: headers_in_file.index(h) for h in self.headers}

            seen_unique_keys = set()
            to_create = []
            preview_data = []
            has_errors = False

            # We wrap the logic in a transaction block to ensure atomicity
            # but catch exceptions at the outer level to return errorcall.
            with transaction.atomic():
                for i, row in enumerate(rows[1:], start=2):
                    data = {
                        h: (str(row[idx]).strip()
                            if row[idx] is not None
                            else "")
                        for h, idx in col_indices.items()
                    }

                    if not any(data.values()):
                        continue

                    is_valid, error_map = validator_func(
                        data, seen_unique_keys)

                    # Guardar para previsualización con errores por campo
                    row_preview = {**data, "_row": i, "_errors": error_map}
                    preview_data.append(row_preview)

                    if not is_valid:
                        has_errors = True
                        if not dry_run:
                            first_err = next(iter(error_map.values()))
                            return errorcall(
                                f"Fila {i}: {first_err}",
                                status.HTTP_400_BAD_REQUEST,
                            )
                    else:
                        if not dry_run:
                            # Map Excel columns to model fields via
                            # field_mapping
                            model_data = {
                                field: data[excel_col]
                                for excel_col, field in field_mapping.items()
                                if excel_col in data
                            }
                            # Merge validator-injected keys (e.g.,
                            # key_country_id, status_id)
                            # that are NOT original Excel column names
                            for key, val in data.items():
                                if (key not in self.headers and
                                        key not in model_data):
                                    model_data[key] = val

                            model_data.update(
                                {
                                    "key_user_created_id": self.user_id,
                                    "key_user_updated_id": self.user_id,
                                }
                            )
                            if "status_id" not in model_data:
                                model_data["status_id"] = STATUS_ACTIVO

                            to_create.append(self.model(**model_data))

                    unique_val = data.get(self.headers[0])
                    seen_unique_keys.add(unique_val)

                if dry_run:
                    return succescall(
                        {
                            "rows": preview_data,
                            "total": len(preview_data),
                            "has_errors": has_errors,
                        },
                        "Validación completada",
                    )

                if to_create:
                    created_instances = self.model.objects.bulk_create(
                        to_create)
                    if audit_save_fn:
                        for instance in created_instances:
                            audit_save_fn(instance)

            return succescall(
                None,
                f"Se han importado {len(to_create)} registros correctamente",
            )

        except Exception as e:
            # If we are here, the transaction inside 'atomic' has been rolled
            # back
            return errorcall(
                f"Error al procesar el archivo: {str(e)}",
                status.HTTP_400_BAD_REQUEST,
            )
